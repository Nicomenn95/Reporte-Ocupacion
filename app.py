import streamlit as st
import pandas as pd
from datetime import datetime
import unicodedata
import io
from openpyxl.styles import PatternFill, Font

# --- CONFIGURACIÓN VISUAL DE LA PÁGINA ---
st.set_page_config(page_title="Reporte Por Tramo", page_icon="🚍", layout="centered")

# --- DICCIONARIO DE RUTAS ---
RUTAS_TPL = {
    '001': ['Terminal Varoli', 'Terminal Talca', 'Casa Matriz', '2 norte con carretera', 'Cruce Pelarco', 'Cruce San Rafael', 'Cruce Camarico', 'Rancagua Sur', 'Colon San Bernardo', 'Terminal Sur'],
    '002': ['Terminal Sur', 'Colon San Bernardo', 'Cruce Rengo', 'Terminal Varoli', 'Terminal Talca'],
    '003': ['Terminal Varoli', 'Terminal Talca', 'Casa matriz', 'Cruce Varoli', 'Cruce Maule', 'Cruce San Javier Sur', 'Cruce Villa Alegre', 'Cruce Linares (Petrobras)', 'Cruce Longavi', 'Cruce Parral (Hospital)', 'Cruce Copihue', 'Cruce San Gregorio', 'Terminal María Teresa', 'Plaza Chillán Viejo', 'KM 21 (Nueva Aldea)', 'Terminal Collao'],
    '004': ['Terminal Collao', 'El Manzano', 'Enlace Penco', 'KM 21 (Nueva Aldea)', 'KM 21(Nueva Aldea).', 'Terminal María Teresa', 'San Carlos La Virgen', 'Cruce San Javier', 'Terminal Varoli', 'Terminal Talca'],
    '005': ['Terminal Varoli', 'Terminal Talca', 'Casa matriz', 'Cruce Varoli', 'Cruce Maule', 'Cruce San Javier Sur', 'Cruce Villa Alegre', 'Cruce Linares (Petrobras)', 'Cruce Longavi', 'Cruce Parral (Hospital)', 'Cruce Copihue', 'Cruce San Gregorio', 'Terminal María Teresa'],
    '006': ['Terminal María Teresa', 'San Carlos La Virgen', 'Cruce San Javier', 'Terminal Varoli', 'Terminal Talca'],
    '007': ['Terminal Cauquenes', 'San Javier Frente PDI', 'Cruce San Javier', 'Terminal Varoli', 'Terminal Talca', '2 norte con carretera', 'Cruce Pelarco', 'Cruce San Rafael', 'Cruce Camarico', 'Rancagua Sur', 'Colon San Bernardo', 'Terminal Sur'],
    '008': ['Terminal Sur', 'Colon San Bernardo', 'Cruce Rengo', 'Terminal Varoli', 'Terminal Talca', 'Cruce Varoli', 'San Javier Frente PDI', 'Terminal Cauquenes'],
    '009': ['Terminal Varoli', 'Terminal Talca', 'Casa matriz', 'Cruce Varoli', 'Cruce Maule', 'Cruce San Javier Sur', 'Cruce Villa Alegre', 'Cruce Linares (Petrobras)', 'Cruce Longavi', 'Cruce Parral (Hospital)', 'Cruce Copihue', 'Cruce San Gregorio', 'Terminal María Teresa', 'Plaza Chillán Viejo', 'Cruce Bulnes', 'Cruce Cabrero', 'Cruce Laja', 'Cruce Perales', 'Terminal María Teresa'], 
    '010': ['Terminal María Teresa', 'Cruce Perales', 'Cruce Laja', 'Cruce Cabrero', 'Cruce Bulnes', 'Terminal María Teresa', 'San Carlos La Virgen', 'Cruce San Javier', 'Terminal Varoli', 'Terminal Talca'],
    '011': ['Frente Municipalidad', 'Paradero Cesfam', 'Terminal Cauquenes', 'San Javier Frente PDI', 'Cruce San Javier', 'Terminal Varoli', 'Terminal Talca', '2 norte con carretera', 'Cruce Pelarco', 'Cruce San Rafael', 'Cruce Camarico', 'Rancagua Sur', 'Colon San Bernardo', 'Terminal Sur'],
    '012': ['Terminal Sur', 'Colon San Bernardo', 'Cruce Rengo', 'Terminal Varoli', 'Terminal Talca', 'Cruce Varoli', 'San Javier Frente PDI', 'Terminal Cauquenes', 'Paradero Cesfam', 'Frente Municipalidad']
}

# --- FUNCIONES ---
def normalizar_texto(texto):
    if not isinstance(texto, str): return ""
    texto = texto.strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFKD', texto) if not unicodedata.combining(c))

def parse_date_time(date_str, time_str):
    try:
        if isinstance(date_str, datetime): d_str = date_str.strftime('%d-%m-%Y')
        else: d_str = str(date_str).strip().replace('/', '-')
        if isinstance(time_str, datetime): t_str = time_str.strftime('%H:%M')
        else: 
            t_str = str(time_str).strip()
            if len(t_str) == 8: t_str = t_str[:5]
        return pd.to_datetime(f"{d_str} {t_str}", format='%d-%m-%Y %H:%M', errors='coerce', dayfirst=True)
    except: return pd.NaT

def get_stop_order(ruta_str, paradero):
    if not isinstance(ruta_str, str): return 999
    ruta_num = str(ruta_str).split()[0]
    if ruta_num in RUTAS_TPL:
        stops = [normalizar_texto(s) for s in RUTAS_TPL[ruta_num]]
        p_clean = normalizar_texto(paradero)
        if p_clean in stops: return stops.index(p_clean)
        for i, s in enumerate(stops):
            if p_clean in s or s in p_clean: return i
    return 999

# --- INTERFAZ WEB ---
st.title("🚍 Generador de Reporte por Tramos")
st.markdown("Sube tu archivo **Reporte resumen de ventas** original y obtén al instante el reporte de ocupación ordenado.")

archivo_subido = st.file_uploader("Arrastra tu Excel aquí", type=["xlsx"])

if archivo_subido is not None:
    with st.spinner("Procesando el archivo de forma automática..."):
        try:
            xls = pd.ExcelFile(archivo_subido)
            sheet_name = xls.sheet_names[0]
            df = pd.read_excel(xls, sheet_name=sheet_name)
            
            if 'Estado' not in df.columns:
                st.error("❌ El archivo no parece ser el reporte correcto (falta la columna 'Estado').")
            else:
                df_pagados = df[df['Estado'] == 'Pagado'].copy()
                df_pagados = df_pagados.dropna(subset=['Folio'])
                df_pagados['Folio'] = df_pagados['Folio'].astype(int).astype(str)
                
                df_pagados['DateTime_Origen'] = df_pagados.apply(lambda row: parse_date_time(row['Fecha salida'], row['Hora salida']), axis=1)
                df_pagados['DateTime_Destino'] = df_pagados.apply(lambda row: parse_date_time(row['Fecha llegada'], row['Hora llegada']), axis=1)
                
                paradas_list = []
                for _, row in df_pagados.iterrows():
                    folio, ruta = row['Folio'], row['Ruta']
                    fecha_viaje_limpia = str(row['Fecha de viaje']).split(' ')[0]
                    
                    paradas_list.append({
                        'Folio': folio, 'Ruta': ruta,
                        'Ciudad': str(row['Origen (ciudad)']).strip() if pd.notna(row['Origen (ciudad)']) else '',
                        'Paradero': str(row['Origen (parada)']).strip() if pd.notna(row['Origen (parada)']) else '',
                        'DateTime': row['DateTime_Origen'], 
                        'Fecha_Oficial': fecha_viaje_limpia 
                    })
                    paradas_list.append({
                        'Folio': folio, 'Ruta': ruta,
                        'Ciudad': str(row['Destino (ciudad)']).strip() if pd.notna(row['Destino (ciudad)']) else '',
                        'Paradero': str(row['Destino (parada)']).strip() if pd.notna(row['Destino (parada)']) else '',
                        'DateTime': row['DateTime_Destino'],
                        'Fecha_Oficial': fecha_viaje_limpia
                    })

                df_paradas = pd.DataFrame(paradas_list).drop_duplicates(subset=['Folio', 'Ciudad', 'Paradero']).copy()
                df_paradas['Orden_Oficial'] = df_paradas.apply(lambda row: get_stop_order(row['Ruta'], row['Paradero']), axis=1)

                res_dfs = []
                for folio, group in df_paradas.groupby('Folio'):
                    group = group.sort_values(by=['DateTime', 'Orden_Oficial']).reset_index(drop=True)
                    for i, row in group.iterrows():
                        c, p = row['Ciudad'], row['Paradero']
                        sub = df_pagados[(df_pagados['Folio'] == folio) & (df_pagados['Origen (ciudad)'].str.strip() == c) & (df_pagados['Origen (parada)'].str.strip() == p)].shape[0]
                        baj = df_pagados[(df_pagados['Folio'] == folio) & (df_pagados['Destino (ciudad)'].str.strip() == c) & (df_pagados['Destino (parada)'].str.strip() == p)].shape[0]
                        group.at[i, 'Suben'] = sub
                        group.at[i, 'Bajan'] = baj
                        group.at[i, 'Revisar orden'] = 'Sí' if row['Orden_Oficial'] == 999 else ''
                    res_dfs.append(group)
                
                if not res_dfs:
                    st.warning("⚠️ No se encontraron pasajes válidos para procesar.")
                else:
                    df_final = pd.concat(res_dfs, ignore_index=True)
                    
                    df_final['Continúan'] = 0
                    df_final['Total tramo ciudad'] = None 
                    df_final['Total pasajeros viaje'] = None
                    
                    for folio, group in df_final.groupby('Folio', sort=False):
                        idx_list = group.index.tolist()
                        
                        # Variable para acumular los que "Suben" por ciudad
                        suben_acumulado_ciudad = 0
                        # Total de pasajeros de todo el viaje para este folio
                        total_viaje = group['Suben'].sum()
                        
                        # Calculamos los que "Continúan" (matemática clásica interna)
                        total_oculto = 0
                        for j, i in enumerate(idx_list):
                            if j == 0: 
                                df_final.at[i, 'Continúan'] = 0
                            else: 
                                df_final.at[i, 'Continúan'] = total_oculto - df_final.at[i, 'Bajan']
                            
                            total_oculto = df_final.at[i, 'Continúan'] + df_final.at[i, 'Suben']
                            
                            # Acumulamos los que suben en esta ciudad específica
                            suben_acumulado_ciudad += df_final.at[i, 'Suben']
                            
                            # Verificamos si es el fin de la ciudad o del viaje
                            es_ultimo_folio = (j == len(idx_list) - 1)
                            es_ultima_ciudad = False
                            
                            if not es_ultimo_folio:
                                next_idx = idx_list[j+1]
                                if df_final.at[i, 'Ciudad'] != df_final.at[next_idx, 'Ciudad']:
                                    es_ultima_ciudad = True
                            
                            # Si cambia la ciudad o termina el viaje, imprimimos la suma de los que subieron
                            if es_ultimo_folio or es_ultima_ciudad:
                                df_final.at[i, 'Total tramo ciudad'] = suben_acumulado_ciudad
                                suben_acumulado_ciudad = 0 # Reseteamos para la siguiente ciudad
                                
                            # Si es la ultimísima parada, imprimimos el gran total
                            if es_ultimo_folio:
                                df_final.at[i, 'Total pasajeros viaje'] = total_viaje

                    dias_semana = {0: 'LUNES', 1: 'MARTES', 2: 'MIÉRCOLES', 3: 'JUEVES', 4: 'VIERNES', 5: 'SÁBADO', 6: 'DOMINGO'}
                    df_final['Fecha_dt_oficial'] = pd.to_datetime(df_final['Fecha_Oficial'], errors='coerce', dayfirst=True)
                    
                    df_final['Día'] = df_final['Fecha_dt_oficial'].dt.dayofweek.map(dias_semana)
                    df_final['Fecha salida'] = df_final['Fecha_dt_oficial'].dt.strftime('%d/%m/%Y')
                    df_final['Hora'] = df_final['DateTime'].dt.strftime('%H:%M')
                    
                    # Columnas finales a exportar
                    cols = ['Folio', 'Fecha salida', 'Día', 'Ciudad', 'Hora', 'Paradero', 'Bajan', 'Suben', 'Continúan', 'Total tramo ciudad', 'Total pasajeros viaje', 'Revisar orden']
                    df_final = df_final[cols]
                    
                    # --- DISEÑO DEL EXCEL ---
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_final.to_excel(writer, sheet_name='Ocupación', index=False)
                        worksheet = writer.sheets['Ocupación']
                        
                        color_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        color_celeste = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
                        
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                        
                        current_color = color_blanco
                        previous_folio = None
                        
                        for row in range(2, worksheet.max_row + 1):
                            current_folio = worksheet.cell(row=row, column=1).value
                            if previous_folio is not None and current_folio != previous_folio:
                                current_color = color_celeste if current_color == color_blanco else color_blanco
                            
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = current_color
                            previous_folio = current_folio
                            
                        # Ajustar anchos
                        worksheet.column_dimensions['D'].width = 15 
                        worksheet.column_dimensions['F'].width = 30 
                        worksheet.column_dimensions['J'].width = 18 # Total tramo ciudad
                        worksheet.column_dimensions['K'].width = 20 # Total pasajeros viaje
                    
                    st.success("✅ ¡Reporte generado con éxito!")
                    st.write("👀 **Vista previa de los datos:**")
                    st.dataframe(df_final.fillna("")) 
                    
                    st.download_button(
                        label="📥 Descargar Ocupación (Excel)",
                        data=output.getvalue(),
                        file_name=archivo_subido.name.replace('.xlsx', '_ocupacion.xlsx'),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        except Exception as e:
            st.error(f"❌ Ocurrió un error inesperado: {e}")
